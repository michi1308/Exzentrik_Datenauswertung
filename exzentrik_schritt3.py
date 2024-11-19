import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import os

# Hilfsfunktion zur Ausgabe in das Text-Widget
def output_to_widget(text_widget, message):
    text_widget.insert(tk.END, message + "\n")
    text_widget.see(tk.END)
    text_widget.update_idletasks()  # Aktualisiert das Textfeld sofort

def markieren_loeschen_runden(path, text_widget):# Definiere eine rote Füllfarbe für "nachbearbeiten"
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Definiere eine gelbe Füllfarbe für Zahlenwert-Bedingungen
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Lade die Arbeitsmappe und bearbeite sie
    try:
        wb = openpyxl.load_workbook(path)
        sheet = wb.active

        # Durchlaufe alle Zellen im Arbeitsblatt
        for row in sheet.iter_rows():
            for cell in row:
                # Markiere alle Zellen mit "nachbearbeiten" rot
                if cell.value == "nachbearbeiten":
                    cell.fill = red_fill
                # Prüfe, ob der Zellwert eine Zahl ist
                elif isinstance(cell.value, (int, float)):
                    # Runde die Zahl auf zwei Nachkommastellen
                    cell.value = round(cell.value, 2)

        # Gehe durch die Spalten N und O (entspricht Spalten 14 bis 15)
        for row in sheet.iter_rows(min_row=2, min_col=14, max_col=15):
            for cell in row:
                if isinstance(cell.value, str) and "-" in cell.value:  # Prüfen, ob der Zellinhalt das "-" enthält
                    try:
                        # Teile die Werte anhand des Minus und strippe Leerzeichen
                        first_val, second_val = map(str.strip, cell.value.split('-'))

                        first_num = float(first_val)
                        second_num = float(second_val)

                        # Debug-Ausgaben
                        # print(f"Processing cell: {cell.value}, first_num: {first_num}, second_num: {second_num}")

                        # Überprüfen der Bereiche
                        first_in_range = -5 <= first_num <= 5
                        second_in_range = 95 <= second_num <= 105

                        # Wenn einer oder beide Werte außerhalb ihres Bereichs liegen, markiere die Zelle gelb
                        if not first_in_range or not second_in_range:
                            # print(f"Marking cell yellow: {cell.value} (first in range: {first_in_range}, second in range: {second_in_range})")
                            cell.fill = yellow_fill  # Markiere die Zelle gelb

                        # Runde beide Werte auf zwei Nachkommastellen
                        first_num_rounded = round(first_num, 2)
                        second_num_rounded = round(second_num, 2)

                        # Setze die gerundeten Werte zurück in die Zelle (mit Kommas)
                        cell.value = f"{str(first_num_rounded).replace('.', ',')} - {str(second_num_rounded).replace('.', ',')}"

                    except ValueError:
                        # Falls der Wert kein Zahlenwert ist oder eine falsche Formatierung hat, einfach überspringen
                        # print(f"ValueError for cell: {cell.value}, skipping...")
                        pass

        # Zeilen entfernen, bei denen in Spalte D und E "nachbearbeiten" steht
        rows_to_remove = []
        for row in sheet.iter_rows(min_row=2):  # Starte ab Zeile 2, um die Kopfzeile zu überspringen
            if row[3].value == "nachbearbeiten" and row[4].value == "nachbearbeiten":  # Spalte D und E
                rows_to_remove.append(row[0].row)  # Speichere die Zeilennummer

        # Entferne die Zeilen in umgekehrter Reihenfolge
        for row_index in reversed(rows_to_remove):
            sheet.delete_rows(row_index)

        # Speichere die Änderungen in der Datei
        wb.save(path)
        output_to_widget(text_widget, "Alle relevanten Zellen wurden entsprechend den Vorgaben markiert und die unerwünschten Zeilen entfernt.")

    # Stelle sicher, dass die Datei geschlossen wird, auch bei Fehlern
    finally:
        wb.close()
        messagebox.showinfo("Erfolg",
                            f"Die Ergebnistabelle wurde erfolgreich überarbeitet und gespeichert unter: {path}")


# GUI-Setup
def datei_auswaehlen(entry):
    file_selected = filedialog.askopenfilename(filetypes=[("Excel Dateien", "*.xlsx")])
    entry.delete(0, tk.END)
    entry.insert(0, file_selected)

def verarbeitung_starten(entry, text_widget):
    path = entry.get()
    if not os.path.isfile(path):
        messagebox.showerror("Fehler", "Bitte wähle eine gültige Datei aus.")
        return
    text_widget.delete(1.0, tk.END)  # Löscht die Textausgabe
    markieren_loeschen_runden(path, text_widget)

def main():
    # Hauptfenster erstellen
    root = tk.Tk()
    root.title("Exzentrik Datenauswertung Schritt 3")

    # Eingabefeld für Dateipfad
    frame = tk.Frame(root)
    frame.pack(padx=10, pady=10)

    entry_label = tk.Label(frame, text="Bitte die Datei angeben:")
    entry_label.grid(row=0, column=0, sticky="w")

    # Einstellen der Breite des Eingabefeldes
    entry = tk.Entry(frame, width=70)
    entry.grid(row=0, column=1)

    browse_button = tk.Button(frame, text="Durchsuchen", command=lambda: datei_auswaehlen(entry))
    browse_button.grid(row=0, column=2, padx=5)

    # Start-Button
    start_button = tk.Button(frame, text="Starten", command=lambda: verarbeitung_starten(entry, text_output))
    start_button.grid(row=1, column=1, pady=10)

    # Text-Widget für die Ausgaben
    text_output = ScrolledText(root, height=20, width=100)
    text_output.pack(padx=10, pady=10)

    root.mainloop()

# Hauptprogramm starten
if __name__ == "__main__":
    main()

