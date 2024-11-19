import pandas as pd
import traceback
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import os

# Hilfsfunktion zur Ausgabe in das Text-Widget
def output_to_widget(text_widget, message):
    text_widget.insert(tk.END, message + "\n")
    text_widget.see(tk.END)
    text_widget.update_idletasks()  # Aktualisiert das Textfeld sofort

def fehlende_werte_berechnen(path_blatt1, path_blatt2, text_widget):

    try:
        # Lese die Daten aus den Excel-Dokumenten
        blatt1 = pd.read_excel(path_blatt1)
        blatt2 = pd.read_excel(path_blatt2)

        # Auswahl der gewünschten Spalten aus blatt2
        blatt2_selected = blatt2[['Dateiname', 'Max Flexion links', 'Max Extension links', 'Max Flexion rechts', 'Max Extension rechts']]

        # Hänge die ausgewählten Spalten von blatt2 an blatt1 an, basierend auf 'Dateiname'
        blatt1 = pd.merge(blatt1, blatt2_selected, on='Dateiname', how='left')

        # Berechnungen pro Zeile
        for index, row in blatt1.iterrows():
            # Werte extrahieren und in float konvertieren, falls möglich
            try:
                d = float(row['Max Flexion links_x']) if pd.notna(row['Max Flexion links_x']) and row['Max Flexion links_x'] != "" else 0
                e = float(row['Max Flexion rechts_x']) if pd.notna(row['Max Flexion rechts_x']) and row['Max Flexion rechts_x'] != "" else 0
                p = float(row['Max Flexion links_y']) if pd.notna(row['Max Flexion links_y']) and row['Max Flexion links_y'] != "" else 0
                q = float(row['Max Extension links']) if pd.notna(row['Max Extension links']) and row['Max Extension links'] != "" else 0
                r = float(row['Max Flexion rechts_y']) if pd.notna(row['Max Flexion rechts_y']) and row['Max Flexion rechts_y'] != "" else 0
                s = float(row['Max Extension rechts']) if pd.notna(row['Max Extension rechts']) and row['Max Extension rechts'] != "" else 0
            except ValueError as ve:
                continue  # Bei Konvertierungsfehler überspringen

            # Berechnungen der Verhältnisse
            if q != 0:
                blatt1.at[index, 'Verhältnis Flexion exzentrisch/Extension konzentrisch links'] = round((1 - (d / q)) * 100, 2)
            else:
                blatt1.at[index, 'Verhältnis Flexion exzentrisch/Extension konzentrisch links'] = "nachbearbeiten"

            if s != 0:
                blatt1.at[index, 'Verhältnis Flexion exzentrisch/Extension konzentrisch rechts'] = round((1 - (e / s)) * 100, 2)
            else:
                blatt1.at[index, 'Verhältnis Flexion exzentrisch/Extension konzentrisch rechts'] = "nachbearbeiten"

            if d != 0:
                blatt1.at[index, 'Verhältnis Flexion konzentrisch/Flexion exzentrisch links'] = round((1 - (p / d)) * 100, 2)
            else:
                blatt1.at[index, 'Verhältnis Flexion konzentrisch/Flexion exzentrisch links'] = "nachbearbeiten"

            if e != 0:
                blatt1.at[index, 'Verhältnis Flexion konzentrisch/Flexion exzentrisch rechts'] = round((1 - (r / e)) * 100, 2)
            else:
                blatt1.at[index, 'Verhältnis Flexion konzentrisch/Flexion exzentrisch rechts'] = "nachbearbeiten"

        # Unerwünschte Spalten entfernen
        blatt1.drop(columns=['Max Flexion links_y', 'Max Flexion rechts_y', 'Max Extension links', 'Max Extension rechts'], inplace=True)


        # Speichern der aktualisierten blatt1 in die neue Excel-Datei
        blatt1.to_excel(path_blatt1, index=False)

        # Spaltenbreite anpassen und Zahlen runden
        wb = load_workbook(path_blatt1)
        ws = wb.active

        # Runde alle numerischen Werte in der Datei auf zwei Nachkommastellen
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.value = round(cell.value, 2)
        # Spaltenbreite auf 20 setzen
        for col in ws.columns:
            max_length = 20
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length

        wb.save(path_blatt1)
        wb.close()

        output_to_widget(text_widget, "Die Daten wurden erfolgreich in 'Ergebnisse_exzentrisch.xlsx' gespeichert.")

    except Exception as e:
        output_to_widget(text_widget, "Ein Fehler ist aufgetreten: " + str(e))
        output_to_widget(text_widget, "Detaillierter Fehlercode: " + traceback.format_exc())

    finally:
        messagebox.showinfo("Erfolg",
                            f"Die Ergebnistabelle wurde erfolgreich erstellt und gespeichert unter: {path_blatt1}")


# GUI-Setup
def datei_auswaehlen(entry):
    file_selected = filedialog.askopenfilename(filetypes=[("Excel Dateien", "*.xlsx")])
    entry.delete(0, tk.END)
    entry.insert(0, file_selected)

def verabeitung_starten(entry_blatt1, entry_blatt2, text_widget):
    path_blatt1 = entry_blatt1.get()
    path_blatt2 = entry_blatt2.get()

    if not os.path.isfile(path_blatt1) or not os.path.isfile(path_blatt2):
        messagebox.showerror("Fehler", "Bitte gib gültige Dateipfade an.")
        return

    text_widget.delete(1.0, tk.END)  # Löscht die Textausgabe
    fehlende_werte_berechnen(path_blatt1, path_blatt2, text_widget)

def main():
    # Hauptfenster erstellen
    root = tk.Tk()
    root.title("Exzentrik Datenauswertung Schritt 2")

    # Frame für Eingabefelder
    frame = tk.Frame(root)
    frame.pack(padx=10, pady=10)

    entry_label1 = tk.Label(frame, text="Pfad für bisherige Exzentrik-Ergebnistabelle angeben:")
    entry_label1.grid(row=0, column=0, sticky="w")

    # Eingabefeld für Pfad Blatt 1
    entry_blatt1 = tk.Entry(frame, width=70)
    entry_blatt1.grid(row=0, column=1)

    browse_button1 = tk.Button(frame, text="Durchsuchen", command=lambda: datei_auswaehlen(entry_blatt1))
    browse_button1.grid(row=0, column=2, padx=5)

    entry_label2 = tk.Label(frame, text="Pfad für Isokinetik-Ergebnistabelle angeben:")
    entry_label2.grid(row=1, column=0, sticky="w")

    # Eingabefeld für Pfad Blatt 2
    entry_blatt2 = tk.Entry(frame, width=70)
    entry_blatt2.grid(row=1, column=1)

    browse_button2 = tk.Button(frame, text="Durchsuchen", command=lambda: datei_auswaehlen(entry_blatt2))
    browse_button2.grid(row=1, column=2, padx=5)

    # Start-Button
    start_button = tk.Button(frame, text="Starten", command=lambda: verabeitung_starten(entry_blatt1, entry_blatt2, text_output))
    start_button.grid(row=2, column=1, pady=10)

    # Text-Widget für die Ausgaben
    text_output = ScrolledText(root, height=20, width=100)
    text_output.pack(padx=10, pady=10)

    root.mainloop()

# Hauptprogramm starten
if __name__ == "__main__":
    main()

