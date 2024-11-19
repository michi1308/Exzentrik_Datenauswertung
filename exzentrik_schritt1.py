
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import pandas as pd
from openpyxl import load_workbook


# Liste der benötigten Arbeitsblätter
required_sheets = [
    "Isokin_Exz_Kon_60_60_Links",
    "Isokin_Exz_Kon_60_60_Rechts"
]

# Hilfsfunktion zur Ausgabe in das Text-Widget
def output_to_widget(text_widget, message):
    text_widget.insert(tk.END, message + "\n")
    text_widget.see(tk.END)
    text_widget.update_idletasks()  # Aktualisiert das Textfeld sofort

def find_neighboring_peaks_with_plateaus(data, peak_index):
    # Funktion zur Suche nach einem Peak (lokales Maximum oder Minimum) in eine Richtung
    def find_peak_in_direction(data, start_index, step):
        # Schleife in Richtung `step` starten
        for idx in range(start_index, len(data) if step > 0 else -1, step):
            # Bedingungen für Randfälle: Anfang und Ende der Datenreihe
            if idx == 0 or idx == len(data) - 1:
                # Randbedingungen: Wenn am Rand ein Maximum oder Minimum vorliegt, akzeptieren
                if (idx == 0 and data[idx] <= data[idx + 1]) or \
                        (idx == len(data) - 1 and data[idx] <= data[idx - 1]) or \
                        (idx == 0 and data[idx] >= data[idx + 1]) or \
                        (idx == len(data) - 1 and data[idx] >= data[idx - 1]):
                    return data[idx]  # Gebe nur den Wert zurück

            # Bedingungen für Plateau- oder Einzelpeaks (lokales Maximum oder Minimum)
            if (data[idx] > data[idx - 1] and data[idx] >= data[idx + 1]) or \
                    (data[idx] < data[idx - 1] and data[idx] <= data[idx + 1]):
                # Bestätigen, dass es sich um ein Plateau oder einen isolierten Peak handelt
                plateau_start = idx
                plateau_end = idx
                # Falls es ein Plateau gibt, erfasse das gesamte Plateau
                while plateau_end + step >= 0 and plateau_end + step < len(data) and \
                        data[plateau_end + step] == data[idx]:
                    plateau_end += step

                # Gebe den Wert des Plateaus zurück
                return data[plateau_start]

        return None  # Wenn kein Peak gefunden wurde

    # Linken und rechten benachbarten Peak suchen und nur die Werte zurückgeben
    left_peak = find_peak_in_direction(data, peak_index - 1, -1)
    right_peak = find_peak_in_direction(data, peak_index + 1, 1)

    return left_peak, right_peak


def excel_dateien_verarbeiten(file_path, file_name, text_widget):
    """Verarbeitet eine einzelne Excel-Datei und gibt ein Dictionary mit den gewünschten Daten zurück."""
    data = {
        'Dateiname': file_name,
        'Name': "n.a.",
        'ID': "n.a.",
        'Max Flexion links': "nachbearbeiten",
        'Max Flexion rechts': "nachbearbeiten",
        'Seitenunterschied Flexion absolut': "nachbearbeiten",
        'Seitenunterschied Flexion relativ (%)': "nachbearbeiten",
        'Verhältnis Flexion exzentrisch/Extension konzentrisch links': "nachbearbeiten",
        'Verhältnis Flexion exzentrisch/Extension konzentrisch rechts': "nachbearbeiten",
        'Verhältnis Flexion konzentrisch/Flexion exzentrisch links': "nachbearbeiten",
        'Verhältnis Flexion konzentrisch/Flexion exzentrisch rechts': "nachbearbeiten",
        'Winkel maximales Drehmoment links Flexion': "nachbearbeiten",
        'Winkel maximales Drehmoment rechts Flexion': "nachbearbeiten",
        'ROM Flexion links': "nachbearbeiten",
        'ROM Flexion rechts': "nachbearbeiten"
    }

    workbook = None
    try:
        # Lade die Arbeitsmappe
        workbook = load_workbook(file_path, data_only=True)

        # Werte aus dem "Wiederholungen"-Blatt extrahieren
        if "Wiederholungen" in workbook.sheetnames:
            sheet = workbook["Wiederholungen"]
            data['Name'] = sheet["A2"].value if sheet["A2"].value else "n.a."
            data['ID'] = sheet["B2"].value if sheet["B2"].value else "n.a."
        else:
            output_to_widget(text_widget, f"{file_name}: Das Arbeitsblatt 'Wiederholungen' fehlt.")

        # Überprüfe jedes erforderliche Arbeitsblatt und finde die maximalen Werte
        for sheet_name in required_sheets:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Extrahiere die Drehmomentwerte aus Spalte C
                torque_values = [cell.value for cell in sheet['C'] if isinstance(cell.value, (int, float))]
                winkel_values = [cell.value for cell in sheet['B'] if isinstance(cell.value, (int, float))]

                if torque_values:
                    max_torque = max(torque_values)
                    max_index = torque_values.index(max_torque) + 1  # Excel-Indizes beginnen bei 1

                    # Hole den zugehörigen Winkelwert aus Spalte B in derselben Zeile wie der Maximalwert
                    angle_cell = sheet[f"B{max_index + 1}"].value

                    # Speichere die Ergebnisse in das Dictionary
                    if sheet_name == "Isokin_Exz_Kon_60_60_Links":
                        data['Max Flexion links'] = max_torque
                        data['Winkel maximales Drehmoment links Flexion'] = angle_cell
                        data[
                            'Index maximales Drehmoment links'] = max_index + 1  # Zeilenindex speichern (Excel-indiziert)
                    elif sheet_name == "Isokin_Exz_Kon_60_60_Rechts":
                        data['Max Flexion rechts'] = max_torque
                        data['Winkel maximales Drehmoment rechts Flexion'] = angle_cell
                        data[
                            'Index maximales Drehmoment rechts'] = max_index + 1  # Zeilenindex speichern (Excel-indiziert)

                # ROM berechnen
                if winkel_values:
                    # ROM links
                    if sheet_name == "Isokin_Exz_Kon_60_60_Links":
                        left_peak, right_peak = find_neighboring_peaks_with_plateaus(winkel_values, data[
                            'Index maximales Drehmoment links'])
                        if left_peak is not None and right_peak is not None:
                            # Vergleiche die Winkel und formatiere die Ausgabe
                            if left_peak < right_peak:
                                data['ROM Flexion links'] = f"{left_peak} - {right_peak}"
                            else:
                                data['ROM Flexion links'] = f"{right_peak} - {left_peak}"
                        else:
                            data['ROM Flexion links'] = "nachbearbeiten"

                    # ROM rechts
                    elif sheet_name == "Isokin_Exz_Kon_60_60_Rechts":
                        left_peak, right_peak = find_neighboring_peaks_with_plateaus(winkel_values, data[
                            'Index maximales Drehmoment rechts'])
                        if left_peak is not None and right_peak is not None:
                            # Vergleiche die Winkel und formatiere die Ausgabe
                            if left_peak < right_peak:
                                data['ROM Flexion rechts'] = f"{left_peak} - {right_peak}"
                            else:
                                data['ROM Flexion rechts'] = f"{right_peak} - {left_peak}"
                        else:
                            data['ROM Flexion rechts'] = "nachbearbeiten"



            else:
                output_to_widget(text_widget, f"{file_name}: Das Arbeitsblatt '{sheet_name}' fehlt.")

        # Seitenunterschied und Verhältnisse für Flexion
        max_flexion_links = data['Max Flexion links']
        max_flexion_rechts = data['Max Flexion rechts']

        if isinstance(max_flexion_links, (int, float)) and isinstance(max_flexion_rechts, (int, float)):
            seitenunterschied_flexion_absolut = abs(max_flexion_links - max_flexion_rechts)
            min_flexion = min(max_flexion_links, max_flexion_rechts)
            max_flexion = max(max_flexion_links, max_flexion_rechts)
            seitenunterschied_flexion_relativ = round((1 - (min_flexion / max_flexion)) * 100, 2)
        else:
            seitenunterschied_flexion_absolut = "nachbearbeiten"
            seitenunterschied_flexion_relativ = "nachbearbeiten"

        # Hinzufügen der berechneten Werte zum Daten-Dictionary
        data.update({
            'Seitenunterschied Flexion absolut': seitenunterschied_flexion_absolut,
            'Seitenunterschied Flexion relativ (%)': seitenunterschied_flexion_relativ
        })

    except Exception as e:
        output_to_widget(text_widget, f"Fehler beim Verarbeiten der Datei {file_name}: {e}")

    finally:
        # Sicherstellen, dass die Arbeitsmappe geschlossen wird
        if workbook:
            workbook.close()

    return data


# Hauptausführungsblock
def daten_verarbeiten_speichern(entry, text_widget):
    """Alle Dateien im Verzeichnis verarbeiten und Ergebnisse speichern"""

    folder_path = entry.get()  # Holen des Pfads aus dem Entry
    if not os.path.isdir(folder_path):
        messagebox.showerror("Fehler", "Bitte gib einen gültigen Ordnerpfad an.")
        return

    result_data = []
    output_to_widget(text_widget, "Datenverarbeitung gestartet...")

    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(folder_path, file_name)
            result_data.append(excel_dateien_verarbeiten(file_path, file_name, text_widget))

    # Erstelle DataFrame und runde numerische Werte auf 2 Dezimalstellen
    result_df = pd.DataFrame(result_data)

    # Entferne die unerwünschten Spalten
    result_df = result_df.drop(columns=['Index maximales Drehmoment links', 'Index maximales Drehmoment rechts'],
                               errors='ignore')

    result_df = result_df.round(2)  # Rundet alle numerischen Werte auf 2 Nachkommastellen

    # Speichern der Ergebnisse in eine Excel-Datei
    output_file_path = os.path.join(folder_path, "Ergebnisse_exzentrisch.xlsx")
    result_df.to_excel(output_file_path, index=False)

    # Spaltenbreite auf 20 setzen
    workbook = load_workbook(output_file_path)
    try:
        sheet = workbook.active
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 20  # Breite auf 20 setzen
        workbook.save(output_file_path)
    finally:
        workbook.close()  # Sicherstellen, dass die Arbeitsmappe immer geschlossen wird

    output_to_widget(text_widget, f"Die Ergebnisse wurden erfolgreich in {output_file_path} gespeichert.")
    messagebox.showinfo("Erfolg",
                        f"Die Ergebnistabelle wurde erfolgreich erstellt und gespeichert unter: {output_file_path}")


# GUI-Setup
def ordner_auswaehlen(entry):
    folder_selected = filedialog.askdirectory()
    entry.delete(0, tk.END)
    entry.insert(0, folder_selected)

def verabeitung_starten(entry, text_widget):
    folder_path = entry.get()
    if not os.path.isdir(folder_path):
        messagebox.showerror("Fehler", "Bitte geb einen gültigen Ordnerpfad an.")
        return
    text_widget.delete(1.0, tk.END)  # Löscht die Textausgabe
    daten_verarbeiten_speichern(entry, text_widget)

def main():
    # Hauptfenster erstellen
    root = tk.Tk()
    root.title("Exzentrik Datenauswertung Schritt 1")

    # Eingabefeld für Ordnerpfad
    frame = tk.Frame(root)
    frame.pack(padx=10, pady=10)

    entry_label = tk.Label(frame, text="Bitte den Ordnerpfad angeben:")
    entry_label.grid(row=0, column=0, sticky="w")

    # Einstellen der Breite des Eingabefeldes
    entry = tk.Entry(frame, width=70)
    entry.grid(row=0, column=1)

    browse_button = tk.Button(frame, text="Durchsuchen", command=lambda: ordner_auswaehlen(entry))
    browse_button.grid(row=0, column=2, padx=5)

    # Start-Button
    start_button = tk.Button(frame, text="Starten", command=lambda: verabeitung_starten(entry, text_output))
    start_button.grid(row=1, column=1, pady=10)

    # Text-Widget für die Ausgaben
    text_output = ScrolledText(root, height=20, width=100)
    text_output.pack(padx=10, pady=10)

    root.mainloop()

# Hauptprogramm starten
if __name__ == "__main__":
    main()


